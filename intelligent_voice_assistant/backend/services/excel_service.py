import re
from typing import Any

import openpyxl

from config import EXCEL_FILE
from services.session_service import get_current_column


def _normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def read_students() -> list[dict[str, Any]]:
    """Read the Excel workbook and return student rows using the actual workbook headers."""
    workbook = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    sheet = workbook.active

    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    students: list[dict[str, Any]] = []

    for row in range(2, sheet.max_row + 1):
        values = [sheet.cell(row=row, column=col).value for col in range(1, len(headers) + 1)]
        row_data: dict[str, Any] = {}

        for header, value in zip(headers, values):
            row_data[header] = value

        if any(v is not None for v in row_data.values()):
            student = {
                "rollNo": row_data.get("Roll_No") or row_data.get("Roll No") or row_data.get("roll_no"),
                "student_name": row_data.get("Name") or row_data.get("Student_Name") or row_data.get("Student Name") or "Unknown",
            }
            for header in headers[1:]:
                student[header] = row_data.get(header, 0) or 0
            students.append(student)

    workbook.close()
    return students


def _resolve_column_number(sheet, requested_column: str):
    requested = _normalize_header(requested_column)
    for cell in sheet[1]:
        if cell.value is None:
            continue
        header_name = str(cell.value).strip()
        header_norm = _normalize_header(header_name)
        if header_norm == requested:
            return cell.column
        if requested in {"assignment", "test", "midterm", "final", "finalterm"}:
            aliases = {
                "assignment": {"assignment"},
                "test": {"test", "quiz", "quizzes"},
                "midterm": {"midterm", "mid"},
                "final": {"final", "finalterm", "finals"},
                "finalterm": {"final", "finalterm", "finals"},
            }
            if header_norm in aliases.get(requested, set()):
                return cell.column
    return None


def update_marks(roll_no: int, marks: int):
    """Update a student's marks in the selected Excel column."""
    workbook = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
    sheet = workbook.active

    current_column = get_current_column()
    if current_column is None:
        workbook.close()
        return False, "No column selected."

    column_number = _resolve_column_number(sheet, current_column)
    if column_number is None:
        workbook.close()
        return False, "Column not found."

    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == roll_no:
            sheet.cell(row=row, column=column_number).value = marks
            workbook.save(EXCEL_FILE)
            workbook.close()
            return True, "Marks Updated Successfully."

    workbook.close()
    return False, "Roll Number not found."